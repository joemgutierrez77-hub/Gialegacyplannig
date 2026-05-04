"""
Excel export — pulls live data from Airtable (or local JSON fallback) and
writes a formatted .xlsx workbook with four sheets:

  1. Pending Apps      — all submitted applications
  2. Issued Policies   — all issued policies with financials
  3. Agent Summary     — per-agent production totals
  4. Agency P&L        — monthly override income vs. chargebacks

Uses openpyxl directly; no pandas dependency needed.
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from config.settings import AGENCY, USE_AIRTABLE

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
_TEAL   = "1A7A6E"   # GIA brand teal
_WHITE  = "FFFFFF"
_GRAY   = "F2F2F2"
_RED    = "C0392B"
_YELLOW = "F39C12"
_GREEN  = "27AE60"

_HEADER_FONT  = Font(bold=True, color=_WHITE, name="Calibri", size=11)
_TITLE_FONT   = Font(bold=True, color=_TEAL,  name="Calibri", size=13)
_NORMAL_FONT  = Font(name="Calibri", size=10)
_HEADER_FILL  = PatternFill("solid", fgColor=_TEAL)
_ALT_FILL     = PatternFill("solid", fgColor=_GRAY)
_THIN         = Side(style="thin", color="CCCCCC")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER       = Alignment(horizontal="center", vertical="center")
_LEFT         = Alignment(horizontal="left",   vertical="center")


def _style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER


def _style_data_row(ws, row: int, n_cols: int, alt: bool = False) -> None:
    fill = _ALT_FILL if alt else PatternFill()
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = _NORMAL_FONT
        cell.fill      = fill
        cell.alignment = _LEFT
        cell.border    = _BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            max(min_width, min(length + 3, max_width))


def _title_row(ws, title: str, n_cols: int, row: int = 1) -> None:
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = _TITLE_FONT
    cell.alignment = _CENTER


def _fmt_currency(val) -> str:
    try:
        return f"${float(val):,.2f}"
    except (TypeError, ValueError):
        return "$0.00"


def _fmt_pct(val) -> str:
    try:
        f = float(val)
        return f"{f * 100:.1f}%" if f <= 1 else f"{f:.1f}%"
    except (TypeError, ValueError):
        return "—"


# ---------------------------------------------------------------------------
# Data loaders — Airtable when active, local JSON otherwise
# ---------------------------------------------------------------------------

def _load_pending() -> list:
    if USE_AIRTABLE:
        from src.airtable_adapter import get_pending_apps
        return get_pending_apps()
    # Local JSON fallback — return empty if no data yet
    import json
    from config.settings import DATA_DIR
    path = Path(DATA_DIR) / "policies" / "pending.json"
    return json.loads(path.read_text()) if path.exists() else []


def _load_issued() -> list:
    if USE_AIRTABLE:
        from src.airtable_adapter import get_issued_policies
        return get_issued_policies()
    from src.modules.profitability import _load_ledger
    return _load_ledger()


def _load_agents() -> list:
    import json
    from config.settings import DATA_DIR
    path = Path(DATA_DIR) / "agents" / "roster.json"
    return json.loads(path.read_text()) if path.exists() else []


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_pending(wb: openpyxl.Workbook, pending: list) -> None:
    ws = wb.create_sheet("Pending Applications")
    headers = [
        "Submit Date", "Agent Name", "Applicant Name",
        "Carrier", "Annual Premium", "Status", "Policy #",
    ]
    n = len(headers)
    _title_row(ws, f"{AGENCY['name']} — Pending Applications", n, row=1)
    ws.append(headers)
    _style_header_row(ws, 2, n)

    for i, app in enumerate(pending, start=3):
        ws.append([
            app.get("submit_date", ""),
            app.get("agent_name",  ""),
            app.get("applicant_name", ""),
            app.get("carrier",     ""),
            app.get("annual_premium", 0),
            app.get("status",      "Pending"),
            app.get("policy_number", ""),
        ])
        _style_data_row(ws, i, n, alt=(i % 2 == 0))
        # Colour-code Status column (col 6)
        status = str(app.get("status", "")).lower()
        cell   = ws.cell(row=i, column=6)
        if status == "approved":
            cell.font = Font(bold=True, color=_GREEN, name="Calibri", size=10)
        elif status in ("declined", "not taken"):
            cell.font = Font(bold=True, color=_RED,   name="Calibri", size=10)
        # Currency format for Annual Premium (col 5)
        ws.cell(row=i, column=5).number_format = '"$"#,##0.00'

    # Totals row
    total_row = len(pending) + 3
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True, name="Calibri", size=10)
    ws.cell(total_row, 5, sum(a.get("annual_premium", 0) for a in pending))
    ws.cell(total_row, 5).number_format = '"$"#,##0.00'
    ws.cell(total_row, 5).font = Font(bold=True, name="Calibri", size=10)

    ws.freeze_panes = "A3"
    _auto_width(ws)


def _sheet_issued(wb: openpyxl.Workbook, issued: list) -> None:
    ws = wb.create_sheet("Issued Policies")
    headers = [
        "Issue Date", "Agent Name", "Policy #", "Carrier",
        "Annual Premium", "Agent Comm %", "Gross Commission",
        "Agency Override", "CB Reserve", "Net to Agent",
        "Persistency", "Status",
    ]
    n = len(headers)
    _title_row(ws, f"{AGENCY['name']} — Issued Policies", n, row=1)
    ws.append(headers)
    _style_header_row(ws, 2, n)

    for i, p in enumerate(issued, start=3):
        ws.append([
            p.get("issue_date",           ""),
            p.get("agent_name",           ""),
            p.get("policy_number",        ""),
            p.get("carrier",              ""),
            p.get("annual_premium",       0),
            p.get("agent_commission_pct", 0),
            p.get("gross_commission",     0),
            p.get("agency_override",      0),
            p.get("chargeback_reserve",   0),
            p.get("net_to_agent",         0),
            p.get("persistency",          ""),
            p.get("status",               "active").title(),
        ])
        _style_data_row(ws, i, n, alt=(i % 2 == 0))

        # Number formats
        for col in (5, 7, 8, 9, 10):
            ws.cell(i, col).number_format = '"$"#,##0.00'
        ws.cell(i, 6).number_format = '0%'

        # Persistency colour coding (col 11)
        pers = p.get("persistency")
        if pers is not None:
            try:
                pf = float(pers)
                colour = _GREEN if pf >= AGENCY["min_persistency_rate"] else _RED
                ws.cell(i, 11).font = Font(bold=True, color=colour,
                                           name="Calibri", size=10)
            except (TypeError, ValueError):
                pass

        # Status colour (col 12)
        status = str(p.get("status", "")).lower()
        if status == "lapsed":
            ws.cell(i, 12).font = Font(bold=True, color=_RED,    name="Calibri", size=10)
        elif status == "active":
            ws.cell(i, 12).font = Font(bold=True, color=_GREEN,  name="Calibri", size=10)

    # Totals
    total_row = len(issued) + 3
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True, name="Calibri", size=10)
    for col_idx, key in enumerate(
        ["annual_premium", "gross_commission", "agency_override",
         "chargeback_reserve", "net_to_agent"],
        start=5
    ):
        val = sum(p.get(key, 0) or 0 for p in issued)
        cell = ws.cell(total_row, col_idx, val)
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(bold=True, name="Calibri", size=10)

    ws.freeze_panes = "A3"
    _auto_width(ws)


def _sheet_agent_summary(wb: openpyxl.Workbook, issued: list, agents: list) -> None:
    ws = wb.create_sheet("Agent Summary")
    headers = [
        "Agent Name", "Active Policies", "Lapsed Policies",
        "Total APV", "Gross Commission", "Agency Override",
        "Chargebacks", "Net Commission", "Vs Target APV",
    ]
    n = len(headers)
    _title_row(ws, f"{AGENCY['name']} — Agent Summary", n, row=1)
    ws.append(headers)
    _style_header_row(ws, 2, n)

    # Aggregate issued policies by agent name
    summary: dict = {}
    for p in issued:
        name = p.get("agent_name", "Unknown")
        if name not in summary:
            summary[name] = {
                "active": 0, "lapsed": 0, "apv": 0.0,
                "gross": 0.0, "override": 0.0, "chargebacks": 0.0, "net": 0.0,
            }
        s = summary[name]
        if str(p.get("status", "")).lower() == "active":
            s["active"]   += 1
            s["apv"]      += p.get("annual_premium",   0) or 0
            s["gross"]    += p.get("gross_commission", 0) or 0
            s["override"] += p.get("agency_override",  0) or 0
            s["net"]      += p.get("net_to_agent",     0) or 0
        else:
            s["lapsed"]     += 1
            cb = p.get("chargeback_actual", p.get("chargeback_reserve", 0)) or 0
            s["chargebacks"]+= cb
            s["net"]        -= cb

    target_apv = AGENCY["target_apv_per_month"]

    for i, (name, s) in enumerate(sorted(summary.items()), start=3):
        vs_target = s["apv"] / target_apv if target_apv else 0
        ws.append([
            name,
            s["active"],
            s["lapsed"],
            s["apv"],
            s["gross"],
            s["override"],
            s["chargebacks"],
            s["net"],
            vs_target,
        ])
        _style_data_row(ws, i, n, alt=(i % 2 == 0))

        for col in (4, 5, 6, 7, 8):
            ws.cell(i, col).number_format = '"$"#,##0.00'
        ws.cell(i, 9).number_format = '0%'

        # Colour vs-target column
        cell = ws.cell(i, 9)
        if vs_target >= 1.0:
            cell.font = Font(bold=True, color=_GREEN,  name="Calibri", size=10)
        elif vs_target >= 0.75:
            cell.font = Font(bold=True, color=_YELLOW, name="Calibri", size=10)
        else:
            cell.font = Font(bold=True, color=_RED,    name="Calibri", size=10)

    ws.freeze_panes = "A3"
    _auto_width(ws)


def _sheet_pnl(wb: openpyxl.Workbook, issued: list) -> None:
    ws = wb.create_sheet("Agency P&L")
    headers = ["Month", "Policies Issued", "Total APV",
               "Override Income", "Chargebacks", "Net Override", "Margin %"]
    n = len(headers)
    _title_row(ws, f"{AGENCY['name']} — Monthly P&L", n, row=1)
    ws.append(headers)
    _style_header_row(ws, 2, n)

    # Group by issue month
    monthly: dict = {}
    for p in issued:
        month = str(p.get("issue_date", ""))[:7]  # "YYYY-MM"
        if not month:
            continue
        if month not in monthly:
            monthly[month] = {
                "count": 0, "apv": 0.0, "override": 0.0, "chargebacks": 0.0
            }
        m = monthly[month]
        m["count"]      += 1
        m["apv"]        += p.get("annual_premium",  0) or 0
        m["override"]   += p.get("agency_override", 0) or 0
        status = str(p.get("status", "")).lower()
        if status == "lapsed":
            m["chargebacks"] += p.get("chargeback_actual",
                                      p.get("chargeback_reserve", 0)) or 0

    for i, (month, m) in enumerate(sorted(monthly.items()), start=3):
        net    = m["override"] - m["chargebacks"]
        margin = net / m["apv"] if m["apv"] else 0
        ws.append([month, m["count"], m["apv"], m["override"],
                   m["chargebacks"], net, margin])
        _style_data_row(ws, i, n, alt=(i % 2 == 0))
        for col in (3, 4, 5, 6):
            ws.cell(i, col).number_format = '"$"#,##0.00'
        ws.cell(i, 7).number_format = '0.0%'

        cell = ws.cell(i, 7)
        if margin >= AGENCY["min_profit_margin"]:
            cell.font = Font(bold=True, color=_GREEN,  name="Calibri", size=10)
        else:
            cell.font = Font(bold=True, color=_RED,    name="Calibri", size=10)

    ws.freeze_panes = "A3"
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_to_excel(output_path: Optional[str] = None) -> str:
    """
    Pull all data and write the full agency tracking workbook.
    Returns the path of the created file.
    """
    if output_path is None:
        stamp       = datetime.today().strftime("%Y-%m-%d")
        output_path = f"GIA_Legacy_Tracker_{stamp}.xlsx"

    pending = _load_pending()
    issued  = _load_issued()
    agents  = _load_agents()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    _sheet_pending(wb, pending)
    _sheet_issued(wb, issued)
    _sheet_agent_summary(wb, issued, agents)
    _sheet_pnl(wb, issued)

    wb.save(output_path)
    return output_path
