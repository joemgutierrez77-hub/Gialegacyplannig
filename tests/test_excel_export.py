"""
Test Excel export — verifies sheet names, headers, and financial totals.
No Airtable or Claude calls are made.
"""
import pytest
import openpyxl
from unittest.mock import patch


SAMPLE_PENDING = [
    {"submit_date": "2026-04-01", "agent_name": "Tom Jones",
     "applicant_name": "Maria G.", "carrier": "Foresters",
     "annual_premium": 1800.0, "status": "Pending", "policy_number": ""},
    {"submit_date": "2026-04-05", "agent_name": "Sara Lee",
     "applicant_name": "John D.",  "carrier": "Mutual of Omaha",
     "annual_premium": 2400.0, "status": "Approved", "policy_number": "POL-001"},
]

SAMPLE_ISSUED = [
    {"issue_date": "2026-04-10", "agent_name": "Tom Jones",
     "policy_number": "POL-002",  "carrier": "Foresters",
     "annual_premium": 3000.0, "agent_commission_pct": 0.70,
     "gross_commission": 2100.0, "agency_override": 150.0,
     "chargeback_reserve": 210.0, "net_to_agent": 1890.0,
     "persistency": 0.91, "status": "active"},
    {"issue_date": "2026-04-15", "agent_name": "Sara Lee",
     "policy_number": "POL-003",  "carrier": "Mutual of Omaha",
     "annual_premium": 4800.0, "agent_commission_pct": 0.70,
     "gross_commission": 3360.0, "agency_override": 240.0,
     "chargeback_reserve": 336.0, "net_to_agent": 3024.0,
     "persistency": 0.78, "status": "active"},
]


@pytest.fixture()
def export_path(tmp_path):
    """Run the export with mocked data and return the workbook."""
    out = str(tmp_path / "test_export.xlsx")
    with patch("src.excel_export._load_pending", return_value=SAMPLE_PENDING), \
         patch("src.excel_export._load_issued",  return_value=SAMPLE_ISSUED), \
         patch("src.excel_export._load_agents",  return_value=[]):
        from src.excel_export import export_to_excel
        export_to_excel(output_path=out)
    return out


def test_workbook_has_four_sheets(export_path):
    wb = openpyxl.load_workbook(export_path)
    assert set(wb.sheetnames) == {
        "Pending Applications", "Issued Policies",
        "Agent Summary", "Agency P&L",
    }


def test_pending_sheet_row_count(export_path):
    ws = openpyxl.load_workbook(export_path)["Pending Applications"]
    # Row 1 = title, row 2 = headers, rows 3+ = data, last row = totals
    data_rows = [r for r in ws.iter_rows(min_row=3, values_only=True)
                 if r[0] not in (None, "TOTAL")]
    assert len(data_rows) == len(SAMPLE_PENDING)


def test_issued_sheet_has_all_policies(export_path):
    ws = openpyxl.load_workbook(export_path)["Issued Policies"]
    policy_numbers = [
        ws.cell(row=r, column=3).value
        for r in range(3, ws.max_row + 1)
        if ws.cell(row=r, column=3).value not in (None, "Policy #")
    ]
    for p in SAMPLE_ISSUED:
        assert p["policy_number"] in policy_numbers


def test_agent_summary_totals(export_path):
    ws = openpyxl.load_workbook(export_path)["Agent Summary"]
    # Collect all APV values from data rows
    apv_values = [
        ws.cell(row=r, column=4).value
        for r in range(3, ws.max_row + 1)
        if ws.cell(row=r, column=4).value not in (None, "Total APV")
    ]
    total_apv = sum(v for v in apv_values if isinstance(v, (int, float)))
    expected  = sum(p["annual_premium"] for p in SAMPLE_ISSUED)
    assert abs(total_apv - expected) < 0.01


def test_pnl_sheet_groups_by_month(export_path):
    ws = openpyxl.load_workbook(export_path)["Agency P&L"]
    months = [
        ws.cell(row=r, column=1).value
        for r in range(3, ws.max_row + 1)
        if ws.cell(row=r, column=1).value not in (None, "Month")
    ]
    assert "2026-04" in months


def test_export_file_is_valid_xlsx(export_path):
    wb = openpyxl.load_workbook(export_path)
    assert len(wb.sheetnames) > 0
